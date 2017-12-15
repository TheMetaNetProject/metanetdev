#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
.. module:: augmentevals
   :platform: Unix
   :synopsis: Used to insert sentence text into IARPA's evaluation excel documents.

IARPA's evaluation excel files don't have the sentence text in them.  Consequently,
it's really hard to figure out what's going on.  This script pulls sentences out
of the input XML files and creates copies of the excel file that has the sentence
text.

.. moduleauthor:: Jisup Hong <jhong@icsi.berkeley.edu>

"""

import sys, os, codecs, subprocess, re
import iarpaxml as ix
import logging
import argparse
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from openpyxl.style import Color, Fill

logger = logging.getLogger(__name__)

def main():
    """
    Runs LM to concept mapping.
    """
    # ------------------------------------------------------------------- #
    # INITIALIZATION
    parser = argparse.ArgumentParser(
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        description="Insert sentence text into eval answers files")
    parser.add_argument('evalfile',
                        help="Excel file containing IARPA's evaluation of our scripts")
    parser.add_argument('detecttestset',
                        help="XML file containing detection testset")
    parser.add_argument('mappingtestset',
                        help='XML file containing mapping testset')
    cmdline = parser.parse_args()
    
    # load the input excel file
    wb = load_workbook(cmdline.evalfile)
    owb = Workbook()
    
    
    # process m4detect
    ows = owb.active
    ows.title = "Detection"
    
    tsxml = ix.TestSetXML('metad', cmdline.detecttestset)
    jdata = tsxml.getJSON()
    # create sentence lookup from jdata
    textLookup = {}
    for sent in jdata['sentences']:
        _, setn, sentn = sent['id'].split(':')
        outn = str(int(setn[2:]))
        textLookup[outn+'.'+sentn] = sent['text']
    
    ws = wb.get_sheet_by_name("Detection")
    rnum = -1
    for row in ws.rows:
        rnum += 1
        (outn,sentn,aktarget,otarget,aksource,osource,mscore,lmbad,goodlm) = (cell.value for cell in row[0:9])
        if isinstance(outn, int):
            text = textLookup[str(outn)+'.'+str(sentn)]
        elif outn and outn.startswith('Output'):
            text='Text'
        else:
            text =''
        print 'mscore is', mscore, 'and has type', type(mscore)
        mscorev = None
        if (isinstance(mscore, str) or isinstance(mscore, unicode)) and mscore.startswith('=IF'):
            print 'mscore is a string, it is', mscore
            mscore=u'=IF(F%d=G%d,1,0)' % (rnum+1,rnum+1)
            mscorev = 1 if aksource==osource else 0
        elif mscore is not None:
            print 'mscore is not a string, it is', mscore
            try:
                mscorev = int(mscore)
            except ValueError:
                pass
        if osource=='# Correct Responses':
            mscore=u'=SUM(H2:H%d)' % (rnum-1)
            lmbad=u'=SUM(I2:I%d)' % (rnum-1)
            goodlm=u'=SUM(J2:J%d)' % (rnum-1)
        ows.append([outn,sentn,text,aktarget,otarget,aksource,osource,mscore,lmbad,goodlm])
        if (mscorev is not None) and (mscorev != 1) and (not goodlm):
            for colnum in xrange(10):
                mycell = ows.cell(row=rnum,column=colnum)
                mycell.style.fill.fill_type = Fill.FILL_SOLID
                mycell.style.fill.start_color.index = Color.DARKRED
                mycell.style.font.color.index = Color.WHITE
    # process m4mapping
    ows = owb.create_sheet()
    ows.title = "Mapping"
    tsxml = ix.TestSetXML('metam', cmdline.mappingtestset)
    jdata = tsxml.getJSON()
    # create sentence lookup from data
    textLookup = {}
    for sent in jdata['sentences']:
        _, setn = sent['id'].split(':')
        outn = str(int(setn[2:]))
        textLookup[outn] = sent['text']
        
    ws = wb.get_sheet_by_name("Mapping")
    rnum = -1
    for row in ws.rows:
        rnum += 1
        (outn,target,source,aktc,otc,tcscore,akcmsource,ocmsource,scorecm) = (cell.value for cell in row[0:9])
        if isinstance(outn, int):
            text = textLookup[str(outn)]
        elif outn and outn.startswith('Output'):
            text='Text'
        else:
            text =''
        scorecmv = None
        if (isinstance(scorecm, str) or isinstance(scorecm, unicode)) and scorecm.startswith('='):
            scorecm=u'=IF(H%d=I%d,1,0)' % (rnum+1,rnum+1)
            print "scorecm is string/unicode",scorecm
            scorecmv = 1 if akcmsource==ocmsource else 0
        elif scorecm is not None:
            print "scorecm is",scorecm
            try:
                scorecmv = 0 if not scorecm else int(scorecm)
            except ValueError:
                pass
        if ocmsource=='# Correct Responses':
            scorecm=u'=SUM(J2:J%d)' % (rnum-1)
        ows.append([outn,text,target,source,aktc,otc,tcscore,akcmsource,ocmsource,scorecm])
        if (outn is not None) and (scorecmv is not None) and (scorecmv != 1):
            for colnum in xrange(10):
                mycell = ows.cell(row=rnum,column=colnum)
                mycell.style.fill.fill_type = Fill.FILL_SOLID
                mycell.style.fill.start_color.index = Color.DARKRED
                mycell.style.font.color.index = Color.WHITE
    fnbase, ext = os.path.splitext(cmdline.evalfile)
    dest_filename = fnbase + '_wsents' + ext
    owb.save(filename = dest_filename)

    
if __name__ == "__main__":
    status = main()
    sys.exit(status)


